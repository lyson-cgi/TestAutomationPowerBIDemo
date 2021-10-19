import logging
import inspect
import pytest
import openpyxl

import os
from dotenv import load_dotenv

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# This BaseClass need to have the knowledge of setup which is define in the conftest.py
@pytest.mark.usefixtures("setup")
class BaseClass:
    def getLogger(self):
        loggerName = inspect.stack()[1][3]
        #logger = logging.getLogger(__name__)
        logger = logging.getLogger(loggerName)

        fileHandler = logging.FileHandler('logfile.log')
        formatter = logging.Formatter("%(asctime)s :%(levelname)s : %(name)s :%(message)s")
        fileHandler.setFormatter(formatter)

        logger.addHandler(fileHandler)  # filehandler object

        logger.setLevel(logging.DEBUG)
        return logger

    def getTestData(self, test_case_id):
        load_dotenv()
        TEST_DATA_FILE_PATH = os.getenv('XLS_TEST_DATA_PATH')
        print("WORKBOOK: ", TEST_DATA_FILE_PATH)
        book = openpyxl.load_workbook(TEST_DATA_FILE_PATH)
        # book = openpyxl.load_workbook("C:\\Users\\Ly-sonLe\\Documents\\Github.com\\TestAutomationPowerBI\\utilities\\test_data.xlsx")
        book.active = 0 # first worksheet
        sheet = book.active
        print("Sheet Title:", sheet.title)
        # print("Test Case ID: ", test_case_id)
        # print("Max Row:", sheet.max_row)
        # print("Max Column:", sheet.max_column)

        Dict = {}

        for i in range(2, sheet.max_row+1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_id:
               _name = sheet.cell(row=i, column=2).value
               _value = sheet.cell(row=i, column=3).value
               _assert_value = sheet.cell(row=i, column=4).value
               # print("asserut_value:", _assert_value)
               Dict[_name] = _value
               Dict['assert_name'] = _assert_value

        return Dict

    def getValue(self, dict, name):
        return dict[name]

    def navToPage(self, xpath):
        elem = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, xpath)))
        elem.click()

    def navToClickable(self, xpath):
        elem = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(
                (By.XPATH, xpath)))
        print("pagesoure inner: ", elem.get_attribute('innerHTML'))
        print("pagesoure outer: ", elem.get_attribute('outerHTML'))
        elem.click()