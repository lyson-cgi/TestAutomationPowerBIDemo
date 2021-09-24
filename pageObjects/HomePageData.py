import openpyxl


class HomePageData:
    test_HomePage_data = [
        {"item1":"Labour Market Outlook Highligts"},
        {"item2": "Region Highligts"},
    ]
    def buildDictionary(self,test_case_id):
        book = openpyxl.load_workbook(
            "C:\\Users\\Ly-sonLe\\Documents\\Github.com\\AutomateTestDemo\\pageObjects\\test_data.xlsx")
        book.active = 0 # first worksheet
        sheet = book.active
        print(sheet.title)
        Dict = {}
        for i in range(2, sheet.max_row+1): # to get rows
            if sheet.cell(row=i, column=1).value == test_case_id:
                for j in range(2, sheet.max_column): # get column
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        cell = sheet.cell(row=2, column=2)
        print("cell: ", cell.value)
        print(Dict)
        return Dict
